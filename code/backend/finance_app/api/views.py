from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.ai.documentintelligence.models import AnalyzeResult, AnalyzeDocumentRequest
from azure.core.credentials import AzureKeyCredential
from azure.storage.blob import (
    BlobServiceClient,
    generate_blob_sas,
    BlobSasPermissions,
    ContentSettings,
)
from django.http import HttpResponse
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.shortcuts import get_object_or_404
from django.conf import settings
from django.contrib.auth import authenticate, logout
from rest_framework import viewsets, filters, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from datetime import datetime, timedelta
from dotenv import load_dotenv
from PIL import Image
import requests
import base64
import os
import time
import openpyxl
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from .models import *
from .serializers import *
import django_filters.rest_framework as filters

load_dotenv()

'''Handles user authentication and management'''
class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):#Return correct serializer based on if a user is being created or not
        if self.action == 'create':
            return UserCreateSerializer
        return UserSerializer

    def get_permissions(self):#Return correct permission based on if a user is being created or not
        if self.action == 'create':
            self.permission_classes = [AllowAny]
        return super().get_permissions()

    def perform_create(self, serializer):#Save new user
        serializer.save()

'''Login viewset using basic authentication'''
class EmailPasswordLoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        email = request.data.get('email')
        password = request.data.get('password')

        if not email or not password:
            return Response({"error": "Email and password are required."}, status=status.HTTP_400_BAD_REQUEST)

        user = authenticate(request, email=email, password=password)

        if user:
            return Response({
                "message": "Login successful!",
                "Authorization": f"Basic {base64.b64encode(f'{email}:{password}'.encode()).decode()}" #encodes email and password for basic authentication
            }, status=status.HTTP_200_OK)

        return Response({"error": "Invalid email or password"}, status=status.HTTP_401_UNAUTHORIZED)
    
'''Logout the user'''
class LogoutView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        logout(request)
        return Response({"message": "Logged out successfully!"}, status=status.HTTP_200_OK)
  
'''Expense viewset'''
class ExpenseViewSet(viewsets.ModelViewSet):
    serializer_class = ExpenseSerializer
    permission_classes = [IsAuthenticated]
    def get_queryset(self): #Returns expenses for authenticated user
        return Expense.objects.filter(user=self.request.user)    
    
    def perform_create(self, serializer): #Assign expense to user
        serializer.save(user=self.request.user)

'''Budget viewset'''
class BudgetViewSet(viewsets.ModelViewSet):
    serializer_class = BudgetSerializer
    permission_classes = [IsAuthenticated]
    def get_queryset(self): #Returns budget for authenticated user
        return Budget.objects.filter(user=self.request.user)    
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_fields = ['id', 'start_date', 'end_date', 'filter_categories']

    def perform_create(self, serializer): #Assign budget to user
        serializer.save(user=self.request.user)

'''Receipt viewset'''
class ReceiptViewSet(viewsets.ModelViewSet):
    serializer_class = ReceiptSerializer
    permission_classes = [IsAuthenticated]
    def get_queryset(self): #Return receipt for authenticated user
        return Receipt.objects.filter(user=self.request.user).order_by('-uploaded_at')
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_fields = ['id','receipt_category']
    
    def perform_create(self, serializer): #Assign receipt to user and budgets
        receipt = serializer.save(user=self.request.user)
        receipt.assign_to_budget()
        
    def update(self, request, *args, **kwargs): #Ensure that changing the category triggers budget updats
        instance = self.get_object()
        old_category = instance.receipt_category  #Store the previous category before update

        response = super().update(request, *args, **kwargs) 

        instance.refresh_from_db()  #Refresh to get updated fields
        new_category = instance.receipt_category  #Get the updated category

        if old_category != new_category:  #Check if the category was changed
            affected_budgets = list(instance.budget.all())
            instance.budget.clear() #Remove from old budget
            instance.assign_to_budget()

            for budget in affected_budgets:  
                budget.update_spending()  #Update old budgets that lost this receipt

            for budget in instance.budget.all():  
                budget.update_spending()  #Update new budgets that gained this receipt
        return response

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        affected_budgets = list(instance.budget.all())  #Store budgets before deletion

        response = super().destroy(request, *args, **kwargs)

        for budget in affected_budgets:
            budget.update_spending()

        return response



'''Viewset that processes receipt images and extracts details about it'''
class ProcessReceiptView(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        
        image_url = request.data.get('image_url')
        uploaded_file = request.FILES.get('image')
        
        if uploaded_file: #Check if there is a uploaded file
            blob_name = generate_filename(uploaded_file.name) #Generate a name
            receiptUrl = upload_image_to_azure(uploaded_file, blob_name) #Turn the uploaded file into a url and store it in blob storage
        elif image_url: #Check if a image url is given instead
            try:
                response = requests.get(image_url, timeout=10) #Download the image from the URL
                if response.status_code != 200:
                    return Response({"error": "Failed to download image from URL."}, status=status.HTTP_400_BAD_REQUEST)

                image_content = ContentFile(response.content) #Convert to Django file
                blob_name = generate_filename("receipt_from_url.jpg") #Generate a name
                receiptUrl = upload_image_to_azure(image_content, blob_name) #Turn the uploaded file into a url and store it in blob storage

            except requests.exceptions.RequestException as e:
                return Response({"error": f"Error fetching image from URL: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)        
        else:
            return Response({"error": "No image file or image_url provided."}, status=status.HTTP_400_BAD_REQUEST)
        
        extracted_data = analyse_receipt(receiptUrl, request) #Process the image and return extracted data from receipt
        return Response(extracted_data, status=status.HTTP_201_CREATED)

'''Create unique name for image'''
def generate_filename(filename):
    timestamp = int(time.time()) #Get current timestamp
    extension = filename.split('.')[-1] #Extract file extension
    return f"{timestamp}.{extension}" #Return unique filename

'''Uploads an image to Azure Blob Storage and returns a URL'''
def upload_image_to_azure(image_file, blob_name):
    
    AZURE_STORAGE_ACCOUNT_NAME = os.getenv("AZURE_STORAGE_ACCOUNT_NAME")
    AZURE_STORAGE_ACCOUNT_KEY = os.getenv("AZURE_STORAGE_ACCOUNT_KEY")
    AZURE_CONTAINER_NAME = os.getenv("AZURE_CONTAINER_NAME")


    compressed_image = compress_image(image_file) #Compress the image
    
    blob_service_client = BlobServiceClient( #Connect to Azure Blob Storage
        f'https://{AZURE_STORAGE_ACCOUNT_NAME}.blob.core.windows.net',
        credential=AZURE_STORAGE_ACCOUNT_KEY
    )

    blob_client = blob_service_client.get_blob_client(container=AZURE_CONTAINER_NAME, blob=blob_name)
    blob_client.upload_blob(compressed_image, overwrite=True)
    blob_client.set_http_headers(content_settings=ContentSettings(content_type="image/png")) #Make sure it returns a Url
    # Upload Image

    
    sas_token = generate_blob_sas( # Generate SAS URL with expiration time which is 1 year in this case
        account_name=AZURE_STORAGE_ACCOUNT_NAME,
        container_name=AZURE_CONTAINER_NAME,
        blob_name=blob_name,
        account_key=AZURE_STORAGE_ACCOUNT_KEY,
        permission=BlobSasPermissions(read=True),
        expiry=datetime.utcnow() + timedelta(days=365 * 100)  # URL expires in 1 year
    )

    sas_url = f"{blob_client.url}?{sas_token}"

    return sas_url

'''Compress the image to reduce file size before uploading'''
def compress_image(image_file):
    
    img = Image.open(image_file)

    img = img.convert("RGB") #Convert to JPEG (reduces file size)

    img_io = BytesIO()
    img.save(img_io, format="JPEG", quality=90) #Adjust quality
    img_io.seek(0)

    return img_io

'''Analyse and extract data from the receipt'''
def analyse_receipt(receiptUrl, request):
    endpoint = str(os.getenv("DOCUMENTINTELLIGENCE_ENDPOINT"))
    key = str(os.getenv("DOCUMENTINTELLIGENCE_API_KEY"))

    document_intelligence_client = DocumentIntelligenceClient(endpoint=endpoint, credential=AzureKeyCredential(key)) #Initialise azure document intelligence client
    if receiptUrl:
        poller = document_intelligence_client.begin_analyze_document(
            "prebuilt-receipt", #se the prebuilt model 'receipt' for scanning and processing
            AnalyzeDocumentRequest(url_source=receiptUrl)
        )
    receipts: AnalyzeResult = poller.result()
    
    if receipts.documents:
        for idx, receipt in enumerate(receipts.documents):
            if receipt.fields:
                merchant_name = receipt.fields.get("MerchantName")
                total = receipt.fields.get("Total")
                items = receipt.fields.get("Items")
                transaction_date_field = receipt.fields.get("TransactionDate")
                receipt_category = receipt.fields.get("ReceiptType")
                receipt_items = []
                if items: #Extract each line from receipt   
                    for idx, item in enumerate(items.get("valueArray")):
                        item_details = {}
                        item_description = item.get("valueObject").get("Description")
                        if item_description:
                            item_details["description"] = {
                            "value": item_description.get("valueString"),
                            }

                        item_quantity = item.get("valueObject").get("Quantity")
                        if item_quantity:
                            item_details["quantity"] = {
                            "value": item_quantity.get("valueString"),
                            }

                        item_total_price = item.get("valueObject").get("TotalPrice")
                        if item_total_price:
                            item_details["total_price"] = {
                            "value": str(item_total_price.get("valueCurrency").get("amount")),
                            }
                            #Categorise the items
                            category = receipt_category.get('valueString')
                            category_choices = {c.value.lower(): c.value for c in CategoryChoices}
                            assigned_category = category_choices.get(category.lower(), CategoryChoices.OTHER)
                            
                        Expense.objects.create(
                            user=request.user,
                            amount=item_total_price.get("valueCurrency").get("amount"),
                            category=assigned_category,
                            date=transaction_date_field.get("valueDate") if transaction_date_field else None,
                            vendor=merchant_name.get('valueString') if merchant_name else "Unknown Merchant",
                        )
                        receipt_items.append(item_details)
    #Categorise the receipts
    category = receipt_category.get('valueString').split(".")[0]
    category_choices = {c.value.lower(): c.value for c in CategoryChoices}
    assigned_category = category_choices.get(category.lower(), CategoryChoices.OTHER)
    
    receipt = Receipt.objects.create(
                                    user=request.user,
                                    image_url=receiptUrl if receiptUrl else None,
                                    merchant=merchant_name.get('valueString') if merchant_name else "Unknown Merchant",
                                    total_amount = float(total.get("valueCurrency", {}).get("amount")) if total else 0.00,
                                    parsed_items=receipt_items,
                                    transaction_date=transaction_date_field.get("valueDate") if transaction_date_field else None,
                                    receipt_category=assigned_category,
                                    )
    receipt.assign_to_budget()
    serializer = ReceiptSerializer(receipt)
    return serializer.data
 
'''Viewset to produce a report of a budget'''   
class BudgetReportView(APIView):  
    permission_classes = [IsAuthenticated]

    def get(self, request, budget_id):
        budget = get_object_or_404(Budget, id=budget_id, user=request.user)
        receipts = Receipt.objects.filter(budget=budget)
        
        #Calculate spending summary
        total_spent = sum(receipt.total_amount or 0 for receipt in receipts)
        category_spending = {}
        total_items = 0
        category_items = {}
        
        for receipt in receipts:
            category = receipt.receipt_category
            category_spending[category] = category_spending.get(category, 0) + (receipt.total_amount or 0)
            
            item_count = len(receipt.parsed_items) if receipt.parsed_items else 1
            total_items += item_count
            category_items[category] = category_items.get(category, 0) + item_count
        
        response_data = {
            "budget": BudgetSerializer(budget).data,
            "total_spent": total_spent,
            "category_spending": category_spending,
            "total_items": total_items,
            "category_items": category_items,
        }
        return Response(response_data, status=status.HTTP_200_OK)

'''Viewset to export a report as a excel file'''
class BudgetReportXlsxView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, budget_id, *args, **kwargs):

        budget = get_object_or_404(Budget, id=budget_id, user=request.user)
        receipts = Receipt.objects.filter(budget=budget)

        category_spending = {}
        total_items = 0
        category_items = {}

        for receipt in receipts:
            category = receipt.receipt_category
            category_spending[category] = category_spending.get(category, 0) + (receipt.total_amount or 0)

            item_count = len(receipt.parsed_items) if receipt.parsed_items else 1
            total_items += item_count
            category_items[category] = category_items.get(category, 0) + item_count

        wb = openpyxl.Workbook() #Create a excel workbook
        bold_font = Font(bold=True)
        
        summary_ws = wb.active #Create budget summart sheet
        summary_ws.title = "Budget Summary"

        summary_headers = ["Budget Name", "User", "Limit Amount", "Total Spent", "Start Date", "End Date"]
        summary_values = [
            budget.name, budget.user.email, f"{budget.limit_amount:.2f}",
            f"{budget.current_spending:.2f}", budget.start_date.strftime('%d-%m-%Y'), budget.end_date.strftime('%d-%m-%Y')
        ]

        for col_num, (header, value) in enumerate(zip(summary_headers, summary_values), start=1):
            summary_ws.cell(row=1, column=col_num, value=header).font = bold_font
            summary_ws.cell(row=2, column=col_num, value=value)
        
        breakdown_ws = wb.create_sheet(title="Spending Breakdown") #Create spending breakdown sheet
        breakdown_ws.append(["Category", "Total Spent", "Item Count"])

        for cell in breakdown_ws[1]:
            cell.font = bold_font
            
        for category, amount in category_spending.items():
            breakdown_ws.append([category, f"{amount:.2f}", category_items.get(category, 0)])

        receipts_ws = wb.create_sheet(title="Receipts") #Create receipts sheet
        receipts_headers = ["ID", "Merchant", "Total Amount", "Transaction Date", "Category", "Item Name", "Item Price"]
        receipts_ws.append(receipts_headers)
        for col_num, header in enumerate(receipts_headers, start=1):
            cell = receipts_ws.cell(row=1, column=col_num, value=header)
            cell.font = bold_font  # Apply bold font

        receipts = receipts.order_by("transaction_date", "uploaded_at")
        for receipt in receipts:
            transaction_date = receipt.transaction_date.strftime('%d/%m/%Y') if receipt.transaction_date else receipt.uploaded_at.strftime('%d/%m/%Y')
            receipt_row = [
                receipt.id,
                receipt.merchant,
                f"{round(receipt.total_amount, 2):.2f}",
                transaction_date,
                receipt.receipt_category
            ]

            if isinstance(receipt.parsed_items, list) and receipt.parsed_items:
                first_item = receipt.parsed_items[0]
                item_name = first_item.get("description", {}).get("value", "Unknown Item")
                item_price = float(first_item.get("total_price", {}).get("value", "0.00"))
                receipt_row.append(item_name)
                receipt_row.append(f"{item_price:.2f}")
            else:
                receipt_row.append("No Items")
                receipt_row.append("-")

            receipts_ws.append(receipt_row)
            
            if isinstance(receipt.parsed_items, list) and len(receipt.parsed_items) > 1: #If there are more than one items
                for item in receipt.parsed_items[1:]:
                    item_name = item.get("description", {}).get("value", "Unknown Item")
                    item_price = float(item.get("total_price", {}).get("value", "0.00"))
                    receipts_ws.append(["", "", "", "", "", item_name, f"{item_price:.2f}"])

        ws = [summary_ws,breakdown_ws,receipts_ws]
        for sheet in ws: #Make the columns of the sheets look nice
            for col_num, col_cells in enumerate(sheet.columns, 1):
                max_length = 0
                col_letter = get_column_letter(col_num)
                for cell in col_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[col_letter].width = adjusted_width
            
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename=budget_{budget_id}_report.xlsx'
        wb.save(response)

        return response