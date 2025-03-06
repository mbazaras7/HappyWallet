from django.test import TestCase
from rest_framework.test import APIClient
from rest_framework import status
from .models import *
from datetime import date
import openpyxl
from io import BytesIO

class UserTests(TestCase):
    def setUp(self): #Create test user
        self.user = User.objects.create_user(email="test@test.com",password="test12345",full_name="test",date_of_birth="2004-09-07")
        self.client = APIClient()

    def test_user_creation(self): #Check if user is created
        self.assertEqual(self.user.email, "test@test.com")

    def test_user_login(self): #Check if user can log in
        response = self.client.post('/login/', {
            'email': 'test@test.com',
            'password': 'test12345'
        })
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('Authorization', response.data) #Check if authorization is returned

class ExpenseTests(TestCase):
    def setUp(self): #Create test user and expense
        self.user = User.objects.create_user(email="test@test.com", password="test12345",full_name="test",date_of_birth="2004-09-07")
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)
        
        self.expense = Expense.objects.create(
            user=self.user,
            name="Water",
            amount=50.00,
            category=CategoryChoices.MEAL.value,
            date=date.today(),
            vendor="Walmart",
        )

    def test_create_expense(self): #Check if expense can be created
        response = self.client.post('/api/expenses/', {
            'name': "Taxi",
            'amount': 20.00,
            'category': CategoryChoices.TRANSPORTATION.value,
            'date': date.today(),
            'vendor': "Uber",
        })
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

    def test_list_expenses(self): #Check if expenses are listed for authenticated user
        response = self.client.get('/api/expenses/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)

class BudgetTests(TestCase):
    def setUp(self): #Create test user and budget
        self.user = User.objects.create_user(email="test@test.com", password="test12345",full_name="test",date_of_birth="2004-09-07")
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

        self.budget = Budget.objects.create(
            user=self.user,
            name="Monthly Budget",
            limit_amount=500.00,
            start_date="2024-02-01",
            end_date="2024-02-28"
        )
        self.budget.filter_categories = [CategoryChoices.MEAL.value, CategoryChoices.ENTERTAINMENT.value]
        self.budget.save()

    def test_create_budget(self): #Check if budget can be created
        response = self.client.post('/api/budgets/', {
            'name': "Test Budget",
            'limit_amount': 100.00,
            'start_date': "2024-02-01",
            'end_date': "2024-02-28"
        })
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

    def test_list_budgets(self): #Check if budgets are listed for authenticated user
        response = self.client.get('/api/budgets/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
            
    def test_budget_filtering_receipts(self): #Check if receipts with the correct category are in the budget
        receipt1 = Receipt.objects.create( #Create allowed receipt
            user=self.user,
            image_url="http://example.com/receipt1.jpg",
            merchant="Supermacs",
            total_amount=20.00,
            transaction_date="2024-02-10",
            receipt_category=CategoryChoices.MEAL.value #In filter categories
        )
        receipt1.assign_to_budget()

        receipt2 = Receipt.objects.create( #Create not allowed receipt
            user=self.user,
            image_url="http://example.com/receipt2.jpg",
            merchant="Pharmacy",
            total_amount=30.00,
            transaction_date="2024-02-15",
            receipt_category=CategoryChoices.HEALTHCARE.value  #Not in filter categories
        )
        receipt2.assign_to_budget()

        self.budget.refresh_from_db() #Refresh budget
        
        self.assertEqual(self.budget.current_spending, 20.00) #Only receipt1 should be counted

class ReceiptTests(TestCase):
    def setUp(self): #Create test user an receipt
        self.user = User.objects.create_user(email="test@test.com", password="test12345",full_name="test",date_of_birth="2004-09-07")
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

        self.receipt = Receipt.objects.create(
            user=self.user,
            image_url="http://example.com/receipt.jpg",
            merchant="McDonald's",
            total_amount=15.99,
            transaction_date="2024-02-01",
            receipt_category=CategoryChoices.MEAL.value,
        )

    def test_create_receipt(self): #Check if receipt can be created
        response = self.client.post('/api/receipts/', {
            'image_url': "http://example.com/new_receipt.jpg",
            'merchant': "Starbucks",
            'total_amount': 5.99,
            'transaction_date': "2024-02-02",
            'receipt_category': CategoryChoices.MEAL.value
        })
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

    def test_list_receipts(self): #Check if receipts are listed for authenticated user
        response = self.client.get('/api/receipts/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_patch_receipt(self): #Check if receipt category can be updated using PATCH   
        patch_data = {"receipt_category": CategoryChoices.HEALTHCARE.value}
        response = self.client.patch(f'/api/receipts/{self.receipt.id}/', patch_data, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.receipt.refresh_from_db()
        self.assertEqual(self.receipt.receipt_category, "Healthcare")

class ProcessReceiptTests(TestCase):
    def setUp(self): #Create test user
        self.user = User.objects.create_user(email="test@test.com", password="test12345",full_name="test",date_of_birth="2004-09-07")
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

    def test_process_receipt(self): #Check if receipt processing works
        response = self.client.post('/api/process-receipt/', {
            'image_url': "https://testcloudblob.blob.core.windows.net/testcloud-blob/1739105642.jpg?se=2125-01-16T12%3A54%3A04Z&sp=r&sv=2025-01-05&sr=b&sig=RgVAeIcPvyryJsN3SpPYhsLkjvxdTu8dz71yYo2yK7Y%3D"
        })
        self.assertIn(response.status_code, [status.HTTP_200_OK, status.HTTP_201_CREATED])
        self.assertEqual(response.data["merchant"], "TESCO\nIRELAND")
        self.assertEqual(response.data["total_amount"], "107.77")
        self.assertEqual(response.data["receipt_category"], "Supplies")
        
class BudgetReportTests(TestCase):
    def setUp(self): #Create test user and budget
        self.user = User.objects.create_user(email="test@test.com", password="test12345",full_name="test",date_of_birth="2004-09-07")
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

        self.budget = Budget.objects.create(
            user=self.user,
            name="Test Budget",
            limit_amount=500.00,
            start_date="2024-02-01",
            end_date="2024-02-28"
        )
        
        self.receipt1 = Receipt.objects.create(
            user=self.user,
            image_url="http://example.com/receipt1.jpg",
            merchant="Nandos",
            total_amount=100.00,
            transaction_date="2024-02-10",
            receipt_category=CategoryChoices.MEAL.value
        )
        self.receipt1.budget.add(self.budget)

        self.receipt2 = Receipt.objects.create(
            user=self.user,
            image_url="http://example.com/receipt2.jpg",
            merchant="Cineworld",
            total_amount=50.00,
            transaction_date="2024-02-15",
            receipt_category=CategoryChoices.ENTERTAINMENT.value
        )
        self.receipt2.budget.add(self.budget)

        self.budget.update_spending() #Refresh budget

    def test_budget_report_view(self): #Check if budget returns correct data
        response = self.client.get(f'/api/budget-report/{self.budget.id}/')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["total_spent"], 150.00)
        self.assertEqual(response.data["category_spending"]["Meal"], 100.00)
        self.assertEqual(response.data["category_spending"]["Entertainment"], 50.00)

class BudgetReportXlsxTests(TestCase):
    def setUp(self): #Create test user and budget for xlsx export
        self.user = User.objects.create_user(email="test@test.com", password="test12345",full_name="test",date_of_birth="2004-09-07")
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

        self.budget = Budget.objects.create(
            user=self.user,
            name="Test Budget XLSX",
            limit_amount=750.00,
            start_date="2024-03-01",
            end_date="2024-03-31"
        )
        
        self.receipt1 = Receipt.objects.create(
            user=self.user,
            image_url="http://example.com/receipt1.jpg",
            merchant="Nandos",
            total_amount=100.00,
            transaction_date="2024-02-10",
            receipt_category=CategoryChoices.MEAL.value
        )
        self.receipt1.budget.add(self.budget)

        self.receipt2 = Receipt.objects.create(
            user=self.user,
            image_url="http://example.com/receipt2.jpg",
            merchant="Cineworld",
            total_amount=50.00,
            transaction_date="2024-02-15",
            receipt_category=CategoryChoices.ENTERTAINMENT.value
        )
        self.receipt2.budget.add(self.budget)

        self.budget.update_spending() #Refresh budget
        
    def test_budget_report_xlsx_view(self): #Check if budget report generates a xlsx file
        response = self.client.get(f'/api/budget-report/{self.budget.id}/xlsx/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        excel_file = BytesIO(response.content)
        workbook = openpyxl.load_workbook(excel_file)
        
        #Check budget summary
        summary_sheet = workbook["Budget Summary"]
        self.assertEqual(summary_sheet.cell(row=2, column=1).value, "Test Budget XLSX")

        #Check spending breakdown
        breakdown_sheet = workbook["Spending Breakdown"]
        self.assertEqual(breakdown_sheet.cell(row=2, column=1).value, "Meal")
        self.assertEqual(float(breakdown_sheet.cell(row=2, column=2).value), 100.00)

        #Check receipts
        receipts_sheet = workbook["Receipts"]
        self.assertEqual(receipts_sheet.cell(row=2, column=2).value, "Nandos")
        self.assertEqual(float(receipts_sheet.cell(row=2, column=3).value), 100.00)
        
class IntegrationTest(TestCase):
    def setUp(self): #Create test user
        self.client = APIClient()
        self.user = User.objects.create_user(email="test@test.com", password="test12345",full_name="test",date_of_birth="2004-09-07")
        self.client.force_authenticate(user=self.user)
        
        self.budget_data = {
            'name': 'Shopping Budget',
            'limit_amount': 500.00,
            'start_date': "2024-03-01",
            'end_date': "2024-03-31",
            'filter_categories': [CategoryChoices.MEAL.value]
        }
        response = self.client.post('/api/budgets/', self.budget_data)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.budget_id = response.data['id']  # Store budget ID for later use

    def test_user_authentication_flow(self): #Check if user registered, can login and can logout
        response = self.client.post('/api/users/', { #Register
            'email': 'test1@test.com',
            'password': 'test12345',
            'full_name': 'test',
            'date_of_birth': '2004-09-07'
        })
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        response = self.client.post('/login/', { #Login
            'email': 'test1@test.com',
            'password': 'test12345'
        })
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('Authorization', response.data)

        response = self.client.post('/logout/') #Logout
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_receipt_processing_and_budget_update(self): #Create a receipt and check if it updates the budget
        receipt_data = {
            'image_url': "http://example.com/receipt.jpg",
            'merchant': "McDonald's",
            'total_amount': 50.00,
            'transaction_date': "2024-03-10",
            'receipt_category': CategoryChoices.MEAL.value  #Matches the budget category
        }
        response = self.client.post('/api/receipts/', receipt_data)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        response = self.client.get('/api/receipts/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)

        response = self.client.get(f'/api/budget-report/{self.budget_id}/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["total_spent"], 50.00)  #Receipt should update budget


    def test_receipt_category_update_affects_budget(self): #Update a receipt category and check if it changes the budget correctly
        receipt_data = {
            'image_url': "http://example.com/receipt.jpg",
            'merchant': "McDonald's",
            'total_amount': 50.00,
            'transaction_date': "2024-03-10",
            'receipt_category': CategoryChoices.MEAL.value
        }
        response = self.client.post('/api/receipts/', receipt_data)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        receipt_id = response.data["id"]

        response = self.client.get(f'/api/budget-report/{self.budget_id}/')
        self.assertEqual(response.data["total_spent"], 50.00)

        patch_data = {"receipt_category": CategoryChoices.HEALTHCARE.value} #Change category to one not in budget filters
        response = self.client.patch(f'/api/receipts/{receipt_id}/', patch_data, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        response = self.client.get(f'/api/budget-report/{self.budget_id}/')
        self.assertEqual(response.data["total_spent"], 0.00)  #Budget should remove the receipt amount
        
    def test_invalid_receipt_does_not_affect_budget(self): #Check if invalid receipt doesnt affect the budget, outside start and end date
        receipt_data = {
            'image_url': "http://example.com/receipt_outside.jpg",
            'merchant': "Old Store",
            'total_amount': 75.00,
            'transaction_date': "2023-12-01",  # BEFORE budget start date
            'receipt_category': CategoryChoices.MEAL.value
        }
        response = self.client.post('/api/receipts/', receipt_data)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        response = self.client.get(f'/api/budget-report/{self.budget_id}/')
        self.assertEqual(response.data["total_spent"], 0.00)  #Receipt should not be counted
        
    def test_delete_receipt_removes_spending_from_budget(self): #Check if current spending is updated when receipt is removed
        receipt_data = {
            'image_url': "http://example.com/receipt.jpg",
            'merchant': "McDonald's",
            'total_amount': 50.00,
            'transaction_date': "2024-03-10",
            'receipt_category': CategoryChoices.MEAL.value
        }
        response = self.client.post('/api/receipts/', receipt_data)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        receipt_id = response.data["id"]

        response = self.client.get(f'/api/budget-report/{self.budget_id}/')
        self.assertEqual(response.data["total_spent"], 50.00)

        response = self.client.delete(f'/api/receipts/{receipt_id}/')
        self.assertEqual(response.status_code, status.HTTP_204_NO_CONTENT)

        response = self.client.get(f'/api/budget-report/{self.budget_id}/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["total_spent"], 0.00)  #Budget should be cleared after deletion
